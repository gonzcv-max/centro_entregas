"""
Gestor de datos del "Centro de Entregas".

Responsable de:
- Saber qué archivos Excel (centros) hay disponibles en data/
- Cargar la hoja "Base Datos Entregas" de un centro como DataFrame limpio
- Guardar los cambios (altas/bajas/cambios) de vuelta al Excel, regenerando
  también las hojas de reportes con valores reales (en vez de tablas
  dinámicas vacías)
- Sincronizar opcionalmente esos cambios a GitHub para que persistan aunque
  Streamlit Cloud reinicie el contenedor (ver utils/github_sync.py)
"""

from __future__ import annotations

import io
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from utils import reports as rpt

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
SHEET_BASE = "Base Datos Entregas"

# Encabezados originales (tal como vienen en el Excel de origen) -> nombre
# interno limpio que usamos en el DataFrame. Se hace strip() a cada
# encabezado leído antes de buscarlo aquí, así que no hace falta poner
# las variantes con espacios de más.
HEADER_TO_INTERNAL = {
    "No. Bol": "no_bol",
    "Producto": "producto",
    "Nombre Comercial": "nombre_comercial",
    "Compartida": "compartida",
    "Fecha Carga": "fecha_carga",
    "Transportista": "transportista",
    "Unidad": "unidad",
    "Capacidad Unidad": "capacidad_unidad",
    "Volumen Cargado": "volumen_cargado",
    "Volumen Descargado": "volumen_descargado",
    "Operador": "operador",
    "Terminal de Carga": "terminal_carga",
    "Sobrante": "sobrante",
    "Merma": "merma",
    "% de Diferencia": "pct_diferencia",
}
INTERNAL_TO_HEADER = {v: k for k, v in HEADER_TO_INTERNAL.items()}

# Orden y encabezados "bonitos" con los que se escribe el Excel de salida.
OUTPUT_COLUMNS = [
    "no_bol",
    "producto",
    "nombre_comercial",
    "compartida",
    "fecha_carga",
    "transportista",
    "unidad",
    "capacidad_unidad",
    "volumen_cargado",
    "volumen_descargado",
    "operador",
    "terminal_carga",
    "diferencia",
    "sobrante",
    "merma",
    "pct_diferencia",
    "estatus",
]

COLUMN_LABELS = {
    "no_bol": "No. Bol",
    "producto": "Producto",
    "nombre_comercial": "Nombre Comercial",
    "compartida": "Compartida",
    "fecha_carga": "Fecha Carga",
    "transportista": "Transportista",
    "unidad": "Unidad",
    "capacidad_unidad": "Capacidad Unidad",
    "volumen_cargado": "Volumen Cargado",
    "volumen_descargado": "Volumen Descargado",
    "operador": "Operador",
    "terminal_carga": "Terminal de Carga",
    "diferencia": "Diferencia (Desc. - Carg.)",
    "sobrante": "Sobrante",
    "merma": "Merma",
    "pct_diferencia": "% de Diferencia",
    "estatus": "Estatus",
}

REQUIRED_FOR_ALTA = [
    "producto",
    "nombre_comercial",
    "fecha_carga",
    "transportista",
    "unidad",
    "capacidad_unidad",
    "volumen_cargado",
]


class CentroError(Exception):
    """Error de negocio al manejar un archivo del centro (no técnico)."""


# --------------------------------------------------------------------------- #
# Listado / alta de archivos del centro
# --------------------------------------------------------------------------- #

def list_centros() -> list[str]:
    """Regresa los nombres de archivo .xlsx disponibles en data/, ordenados."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(p.name for p in DATA_DIR.glob("*.xlsx") if not p.name.startswith("~$"))


def centro_path(filename: str) -> Path:
    return DATA_DIR / filename


def guardar_nuevo_centro(filename: str, file_bytes: bytes) -> str:
    """Valida y guarda un Excel nuevo dentro del centro. Regresa el nombre final."""
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    filename = filename.strip().replace("/", "-").replace("\\", "-")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise CentroError(f"No se pudo abrir el archivo como Excel válido: {exc}") from exc

    if SHEET_BASE not in wb.sheetnames:
        raise CentroError(
            f"El archivo no tiene una hoja llamada '{SHEET_BASE}'. "
            f"Hojas encontradas: {', '.join(wb.sheetnames)}"
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    dest = centro_path(filename)
    if dest.exists():
        stem, suffix = dest.stem, dest.suffix
        dest = dest.with_name(f"{stem}_{datetime.now():%Y%m%d_%H%M%S}{suffix}")

    dest.write_bytes(file_bytes)
    return dest.name


# --------------------------------------------------------------------------- #
# Carga y limpieza de la Base de Datos de Entregas
# --------------------------------------------------------------------------- #

def _to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        # columnas Sobrante/Merma a veces traen texto ("Merma"/"Sobrante")
        # en vez de número: eso se recalcula, no se usa tal cual.
        try:
            return float(value.replace(",", ""))
        except ValueError:
            return None
    return float(value)


def load_centro(filename: str) -> pd.DataFrame:
    """Lee la hoja Base Datos Entregas y regresa un DataFrame limpio y
    recalculado (diferencia, sobrante/merma, % de diferencia, estatus)."""
    path = centro_path(filename)
    if not path.exists():
        raise CentroError(f"El archivo '{filename}' ya no existe en el centro.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_BASE not in wb.sheetnames:
        raise CentroError(f"'{filename}' no tiene la hoja '{SHEET_BASE}'.")
    ws = wb[SHEET_BASE]

    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [(h.strip() if isinstance(h, str) else h) for h in raw_headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)

    # nos quedamos solo con las columnas que reconocemos + renombrado interno
    keep = {h: HEADER_TO_INTERNAL[h] for h in headers if h in HEADER_TO_INTERNAL}
    df = df[list(keep.keys())].rename(columns=keep)

    for col in ["no_bol", "capacidad_unidad", "volumen_cargado", "volumen_descargado"]:
        if col in df.columns:
            df[col] = df[col].apply(_to_number)

    if "fecha_carga" in df.columns:
        df["fecha_carga"] = pd.to_datetime(df["fecha_carga"], errors="coerce")

    for col in ["producto", "nombre_comercial", "compartida", "transportista",
                "unidad", "operador", "terminal_carga"]:
        if col in df.columns:
            df[col] = df[col].astype("object").where(df[col].notna(), None)
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)

    df = recalcular(df)
    df = df.reset_index(drop=True)
    df.insert(0, "id", df.index)
    return df


def recalcular(df: pd.DataFrame) -> pd.DataFrame:
    """Recalcula diferencia, sobrante, merma, % de diferencia y estatus
    a partir de volumen_cargado / volumen_descargado. Se llama tanto al
    cargar el archivo como después de cualquier alta/baja/cambio."""
    df = df.copy()
    cargado = pd.to_numeric(df.get("volumen_cargado"), errors="coerce")
    descargado = pd.to_numeric(df.get("volumen_descargado"), errors="coerce")

    diferencia = descargado - cargado
    df["diferencia"] = diferencia

    df["sobrante"] = diferencia.where(diferencia > 0)
    df["merma"] = diferencia.where(diferencia <= 0)

    with pd.option_context("mode.chained_assignment", None):
        pct = diferencia / cargado
    df["pct_diferencia"] = pct.replace([float("inf"), float("-inf")], pd.NA)

    df["estatus"] = descargado.isna().map({True: "Pendiente de descarga", False: "Entregado"})
    return df


# --------------------------------------------------------------------------- #
# Validación de altas / cambios
# --------------------------------------------------------------------------- #

def validar_dataframe(df: pd.DataFrame) -> list[str]:
    """Errores que SÍ bloquean el guardado: campos obligatorios faltantes o
    volúmenes imposibles (negativos). Deliberadamente no bloquea por exceso
    de capacidad: en la operación real hay cargas por encima de la
    capacidad nominal y no queremos impedir que se guarden datos que ya
    existían en el Excel original (ver advertencias_dataframe)."""
    errores = []
    for i, row in df.iterrows():
        faltantes = [COLUMN_LABELS[c] for c in REQUIRED_FOR_ALTA if pd.isna(row.get(c)) or row.get(c) == ""]
        if faltantes:
            etiqueta = row.get("no_bol") or f"fila {i + 1}"
            errores.append(f"Registro {etiqueta}: falta(n) {', '.join(faltantes)}")

        cargado = row.get("volumen_cargado")
        if pd.notna(cargado) and cargado < 0:
            errores.append(f"Registro {row.get('no_bol', i + 1)}: volumen cargado no puede ser negativo.")
    return errores


def advertencias_dataframe(df: pd.DataFrame) -> list[str]:
    """Avisos que NO bloquean el guardado, solo se muestran como alerta
    (por ejemplo, volumen cargado por encima de la capacidad nominal de
    la unidad, algo que ocurre en la operación real)."""
    avisos = []
    for i, row in df.iterrows():
        cargado = row.get("volumen_cargado")
        capacidad = row.get("capacidad_unidad")
        if pd.notna(cargado) and pd.notna(capacidad) and capacidad and cargado > capacidad * 1.05:
            avisos.append(
                f"Registro {row.get('no_bol', i + 1)}: volumen cargado ({cargado:g}) "
                f"excede la capacidad de la unidad ({capacidad:g})."
            )
    return avisos


# --------------------------------------------------------------------------- #
# Guardado a Excel (Base + hojas de reportes con valores reales)
# --------------------------------------------------------------------------- #

def _autosize(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        largo = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(largo + 2, 10), 45)


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Arma el archivo Excel completo (base + 6 reportes) en memoria."""
    df = recalcular(df)
    export = df[[c for c in OUTPUT_COLUMNS if c in df.columns]].rename(columns=COLUMN_LABELS)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name=SHEET_BASE, index=False)
        _autosize(writer.sheets[SHEET_BASE], export)

        for sheet_name, report_df in rpt.generar_todos_los_reportes(df).items():
            report_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _autosize(writer.sheets[sheet_name], report_df)

    return buffer.getvalue()


def guardar_centro(filename: str, df: pd.DataFrame) -> None:
    """Guarda el DataFrame (ya validado) de vuelta al archivo del centro."""
    data = to_excel_bytes(df)
    centro_path(filename).write_bytes(data)
